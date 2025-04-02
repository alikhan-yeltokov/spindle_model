#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Tue Oct 29 14:56:48 2024

@author: ayn6k
"""
import os
import pandas as pd
import numpy as np
from scipy.optimize import curve_fit
from openpyxl import Workbook, load_workbook

def decaying_exponential(t, A, tau, C):
    """Exponential decay model: y(t) = A * exp(-t / tau) + C"""
    return A * np.exp(-t / tau) + C

# def calculate_characteristic_time(df):
#     """Calculate the characteristic time from the average of all columns."""
#     # Calculate the average of all columns
#     avg = np.mean(df, axis=1).values
#     std = np.std(df, axis=1)
#     # Generate a time array (assuming evenly spaced data points)
#     t = np.arange(len(avg)) * 0.05
#     # Fit the mean data to the decaying exponential function
#     initial_guess = [avg[0], 300.0, avg[-1]]  # A, tau, C
#     params, covariance = curve_fit(decaying_exponential, t, avg, p0=initial_guess)
#     # Extract the fitted parameters
#     A_fitted, tau_fitted, C_fitted = params
#     return tau_fitted



def calculate_characteristic_time(df):
    """
    Calculate the characteristic time and decay value for each column, then return the averages and standard deviations.
    
    Parameters:
    df (pd.DataFrame): Input DataFrame where each column is a time series.
    
    Returns:
    float: Average characteristic time.
    float: Standard deviation of characteristic times.
    float: Average decay value (C).
    float: Standard deviation of decay values (C).
    """
    # Generate a time array (assuming evenly spaced data points)
    t = np.arange(len(df)) * 0.05
    
    # Lists to store characteristic times and decay values
    tau_list = []
    C_list = []
    
    # Loop through each column in the DataFrame
    for column in df.columns:
        # Extract the time series data for the current column
        y = df[column].values
        
        # Fit the decaying exponential function to the current column
        initial_guess = [y[0], 300.0, y[-1]]  # A, tau, C
        params, _ = curve_fit(decaying_exponential, t, y, p0=initial_guess)
        
        # Extract the characteristic time (tau) and decay value (C)
        tau_list.append(params[1])  # tau
        C_list.append(params[2])   # C
    
    # Calculate the average and standard deviation of the characteristic times
    tau_avg = np.mean(tau_list)
    tau_std = np.std(tau_list)
    
    # Calculate the average and standard deviation of the decay values
    C_avg = np.mean(C_list)
    C_std = np.std(C_list)
    
    return tau_avg, tau_std, C_avg, C_std


def write_row_to_excel(file_path, data):
    """Write a row of data to an Excel file."""
    # If the file does not exist, create it with headers
    if not os.path.exists(file_path):
        wb = Workbook()
        ws = wb.active
        ws.title = "Results"
        # Write the header row
        headers = ['Filename', 'tau_avg', 'tau_std', 'C_avg', 'C_std',  'N Pull MT', 'N Push MT']
        ws.append(headers)
        wb.save(file_path)
    
    # Append the data to the file
    wb = load_workbook(file_path)
    ws = wb["Results"]
    ws.append(data)
    wb.save(file_path)

def process_excel_files(folder_path, output_file):
    """Process all Excel files in the given folder and record characteristic times."""
    # Loop through all Excel files in the specified folder
    for filename in os.listdir(folder_path):
        if filename.endswith(".xlsx") or filename.endswith(".xls"):
            file_path = os.path.join(folder_path, filename)
            print(f"Processing file: {file_path}")
            
            # Read the Excel file, skipping the index column
            df_angle = pd.read_excel(file_path, index_col=0, sheet_name='Angle')
            
            # Calculate the characteristic time
            char_time = calculate_characteristic_time(df_angle)
            
            df_pull_n = pd.read_excel(file_path, index_col=0, sheet_name='N_pull')
            df_push_n = pd.read_excel(file_path, index_col=0, sheet_name='N_push')
            
            pull_column_means = df_pull_n.mean()
            push_column_means = df_push_n.mean()

            # Calculate the overall mean
            pull_mean = pull_column_means.mean()
            push_mean = push_column_means.mean()
            
   
            # Write the result directly to the Excel file
            write_row_to_excel(output_file, [filename, char_time, pull_mean, push_mean])
            
def process_excel_files_last(folder_path, output_file):
    """Process all Excel files in the given folder and record characteristic times."""
    # Loop through all Excel files in the specified folder
    for filename in os.listdir(folder_path):
        if filename.endswith(".xlsx") or filename.endswith(".xls"):
            file_path = os.path.join(folder_path, filename)
            print(f"Processing file for last: {file_path}")
            
            # Read the Excel file, skipping the index column
            df_angle = pd.read_excel(file_path, index_col=0, sheet_name='Angle')
            
            # Calculate the characteristic time
            char_time = calculate_characteristic_time(df_angle)
            
            df_pull_n = pd.read_excel(file_path, index_col=0, sheet_name='N_pull')
            df_push_n = pd.read_excel(file_path, index_col=0, sheet_name='N_push')
            
 
            
            pull_last_entries = df_pull_n.iloc[-1]
            push_last_entries = df_push_n.iloc[-1]
            
            # Calculate the mean of the last entries
            pull_mean = pull_last_entries.mean()
            push_mean = push_last_entries.mean()
            
            # Write the result directly to the Excel file
            write_row_to_excel(output_file, [filename, char_time, pull_mean, push_mean])

def process_excel_files_middle(folder_path, output_file):
    """Process all Excel files in the given folder and record characteristic times."""
    # Loop through all Excel files in the specified folder
    for filename in os.listdir(folder_path):
        if filename.endswith(".xlsx") or filename.endswith(".xls"):
            file_path = os.path.join(folder_path, filename)
            print(f"Processing file for mid: {file_path}")
            
            # Read the Excel file, skipping the index column
            df_angle = pd.read_excel(file_path, index_col=0, sheet_name='Angle')
            
            # Calculate the characteristic time
            tau_avg, tau_std, C_avg, C_std = calculate_characteristic_time(df_angle)
            
            df_pull_n = pd.read_excel(file_path, index_col=0, sheet_name='N_pull')
            df_push_n = pd.read_excel(file_path, index_col=0, sheet_name='N_push')
            
 
            pull_row_12000 = df_pull_n.loc[12000] if 12000 in df_pull_n.index else None
            push_row_12000 = df_push_n.loc[12000] if 12000 in df_push_n.index else None
            
            # Calculate the mean of the entries in the row with index 12000
            pull_mean = pull_row_12000.mean() if pull_row_12000 is not None else None
            push_mean = push_row_12000.mean() if push_row_12000 is not None else None
            
            # Write the result directly to the Excel file
            write_row_to_excel(output_file, [filename, tau_avg, tau_std, C_avg, C_std, pull_mean, push_mean])
            
# Usage example
if __name__ == "__main__":
    input_folder = "./char_project"  # Specify your folder path here
    # output_file_last= "characteristic_times_pull_push_count_last_entry.xlsx"  # Specify the output file name
    
    # # process_excel_files_last(input_folder, output_file_last)
    
    # process_excel_files_last(input_folder, output_file_last)
    # print(f"Characteristic times have been recorded in {output_file_last}.")

    output_file_mid= "average_after_characteristic_times_pull_push_count_mid_entry.xlsx"  # Specify the output file name
    process_excel_files_middle(input_folder, output_file_mid)
    print(f"Characteristic times have been recorded in {output_file_mid}.")
# import os
# import pandas as pd
# import numpy as np
# from scipy.optimize import curve_fit

# def decaying_exponential(t, A, tau, C):
#     """Exponential decay model: y(t) = A * exp(-t / tau) + C"""
#     return A * np.exp(-t / tau) + C

# def calculate_characteristic_time(df):
#     """Calculate the characteristic time from the average of all columns."""
#     # Calculate the average of all columns
#     avg = np.mean(df, axis=1).values
#     std =np.std(df, axis=1)
#     # Step 4: Generate a time array (assuming evenly spaced data points)
#     t = np.arange(len(avg))*0.05
#     # Step 5: Fit the mean data to the decaying exponential function
#     initial_guess = [avg[0], 300.0, avg[-1]]  # A, tau, C
#     params, covariance = curve_fit(decaying_exponential, t, avg, p0=initial_guess)
#     # Extract the fitted parameters
#     A_fitted, tau_fitted, C_fitted = params
#     return tau_fitted

# def process_excel_files(folder_path, output_file):
#     """Process all Excel files in the given folder and record characteristic times."""
#     results = []

#     # Loop through all Excel files in the specified folder
#     for filename in os.listdir(folder_path):
#         if filename.endswith(".xlsx") or filename.endswith(".xls"):
#             file_path = os.path.join(folder_path, filename)
#             print(f"Processing file: {file_path}")
            
#             # Read the Excel file, skipping the index column
#             df_angle = pd.read_excel(file_path, index_col=0, sheet_name='Angle')
            
#             # Calculate the characteristic time
#             char_time = calculate_characteristic_time(df_angle)
            
#             df_pull_n = pd.read_excel(file_path, index_col=0, sheet_name='N_pull')
#             df_push_n = pd.read_excel(file_path, index_col=0, sheet_name='N_push')
            
#             pull_column_means = df_pull_n.mean()
#             push_column_means = df_push_n.mean()

#             # Calculate the overall mean
#             pull_mean = pull_column_means.mean()
#             push_mean = push_column_means.mean()
            
#             results.append((filename, char_time, pull_mean, push_mean))
    
#     # Create a DataFrame to store the results
#     results_df = pd.DataFrame(results, columns=['Filename', 'Characteristic Time', 'N Pull MT', 'N Push MT'])
    
#     # Write the results to the output Excel file
#     results_df.to_excel(output_file, index=False)

# # Usage example
# if __name__ == "__main__":
#     # input_folder = "./char_project_pull_n"  # Specify your folder path here
#     input_folder = "./dummy"  # Specify your folder path here
#     output_file = "characteristic_times_pull_push_count.xlsx"  # Specify the output file name
    
#     process_excel_files(input_folder, output_file)
#     print(f"Characteristic times have been recorded in {output_file}.")
